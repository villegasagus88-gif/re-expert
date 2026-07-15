"""
Entregables de análisis financiero (PDF + Excel) para el chat — Capa 2.

Tras un análisis (factibilidad, flujo de fondos, residual, tasación, inversión,
impuestos), el bot puede generar un documento DESCARGABLE y un mensaje listo
para WhatsApp. Esto es lo que un modelo genérico no puede: no solo te asesora,
te entrega el archivo para la reunión/banco.

Diseño:
  - El modelo arma `secciones` con el contenido del análisis (pares dato/valor
    y/o tablas). Pasa los MONTOS como números crudos para que el Excel sea
    editable; etiquetas y % con símbolo van como texto.
  - Renderizamos PDF (reportlab) y Excel (openpyxl) con el mismo contenido.
  - Subimos a Supabase Storage (URL firmada 24h) con fallback a disco, reusando
    los helpers de document_service.
  - Devolvemos los links + un mensaje armado para compartir.

Las dependencias pesadas (reportlab/openpyxl/document_service) se importan de
forma diferida dentro de las funciones para que un fallo de una no tumbe el
import del módulo ni el arranque del backend.
"""
from __future__ import annotations

import asyncio
import io
import logging
import re
import uuid
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)


# ── helpers puros (testeables sin dependencias) ──
def _esc(s: Any) -> str:
    s = "" if s is None else str(s)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _slug(s: str | None) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", (s or "analisis").strip().lower())
    return s.strip("-")[:40] or "analisis"


def _fmt(v: Any) -> str:
    """Formatea números a estilo AR (1.234.567,89) para mostrar en PDF."""
    if isinstance(v, bool):
        return "Sí" if v else "No"
    if isinstance(v, (int, float)):
        s = f"{float(v):,.2f}"
        # de 1,234,567.89 (US) a 1.234.567,89 (AR)
        s = s.replace(",", "§").replace(".", ",").replace("§", ".")
        if s.endswith(",00"):
            s = s[:-3]
        return s
    return str(v)


def _xcell(v: Any) -> Any:
    """Valor para una celda de Excel: números crudos quedan numéricos."""
    if isinstance(v, (int, float, str)):
        return v
    if v is None:
        return ""
    return str(v)


def _norm_secciones(secciones: Any) -> list[dict]:
    out: list[dict] = []
    if not isinstance(secciones, list):
        return out
    for sec in secciones:
        if not isinstance(sec, dict):
            continue
        tipo = (sec.get("tipo") or "kv").strip().lower()
        if tipo not in ("kv", "tabla"):
            tipo = "kv"
        filas = sec.get("filas") if isinstance(sec.get("filas"), list) else []
        out.append(
            {
                "titulo": sec.get("titulo") or "",
                "tipo": tipo,
                "headers": sec.get("headers") if isinstance(sec.get("headers"), list) else [],
                "filas": filas,
            }
        )
    return out


# ── render PDF (reportlab) ──
def render_analisis_pdf(
    titulo: str, secciones: list[dict], resumen: str | None = None, proyecto: str | None = None
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    secciones = _norm_secciones(secciones)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=1.8 * cm, bottomMargin=1.8 * cm, title=(titulo or "Análisis")[:120],
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=colors.HexColor("#0f172a"), fontSize=16)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#1e293b"), fontSize=12, spaceBefore=8)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=14)
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=8, textColor=colors.HexColor("#64748b"))
    cell = ParagraphStyle("cell", parent=styles["BodyText"], fontSize=9, leading=12)
    cell_b = ParagraphStyle("cellb", parent=cell, fontName="Helvetica-Bold", textColor=colors.HexColor("#334155"))

    story: list[Any] = [Paragraph(_esc(titulo or "Análisis"), h1)]
    meta = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — RE Expert"
    if proyecto:
        meta = f"Proyecto: {_esc(proyecto)} · " + meta
    story.append(Paragraph(meta, small))
    if resumen:
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(_esc(resumen), body))
    story.append(Spacer(1, 0.4 * cm))

    head_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ])
    kv_style = TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ])

    for sec in secciones:
        if sec["titulo"]:
            story.append(Paragraph(_esc(sec["titulo"]), h2))
        if sec["tipo"] == "tabla":
            headers = sec["headers"]
            rows: list[list[Any]] = []
            if headers:
                rows.append([Paragraph(_esc(h), cell_b) for h in headers])
            for f in sec["filas"]:
                if not isinstance(f, (list, tuple)):
                    continue
                rows.append([Paragraph(_esc(_fmt(c)), cell) for c in f])
            if rows:
                t = Table(rows, repeatRows=1 if headers else 0)
                t.setStyle(head_style)
                story.append(t)
        else:  # kv
            rows = []
            for f in sec["filas"]:
                if not isinstance(f, (list, tuple)) or len(f) < 1:
                    continue
                k = f[0]
                v = f[1] if len(f) > 1 else ""
                rows.append([Paragraph(_esc(str(k)), cell_b), Paragraph(_esc(_fmt(v)), cell)])
            if rows:
                t = Table(rows, colWidths=[6 * cm, 10.5 * cm])
                t.setStyle(kv_style)
                story.append(t)
        story.append(Spacer(1, 0.3 * cm))

    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        "Documento generado automáticamente por RE Expert. Verificá los datos antes de decidir.",
        small,
    ))
    doc.build(story)
    return buf.getvalue()


# ── render Excel (openpyxl) ──
def render_analisis_xlsx(
    titulo: str, secciones: list[dict], resumen: str | None = None, proyecto: str | None = None
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    secciones = _norm_secciones(secciones)
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis"

    dark = PatternFill("solid", fgColor="0F172A")
    soft = PatternFill("solid", fgColor="F1F5F9")
    white_b = Font(color="FFFFFF", bold=True)
    label_f = Font(bold=True, color="334155")
    wrap = Alignment(wrap_text=True, vertical="top")

    r = 1
    ws.cell(r, 1, titulo or "Análisis").font = Font(bold=True, size=15, color="0F172A")
    r += 1
    meta = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — RE Expert"
    if proyecto:
        meta = f"Proyecto: {proyecto} · " + meta
    ws.cell(r, 1, meta).font = Font(size=9, color="64748B")
    r += 1
    if resumen:
        r += 1
        c = ws.cell(r, 1, resumen)
        c.font = Font(italic=True)
        c.alignment = wrap
        r += 1
    r += 1

    for sec in secciones:
        if sec["titulo"]:
            ws.cell(r, 1, sec["titulo"]).font = Font(bold=True, size=12, color="1E293B")
            r += 1
        if sec["tipo"] == "tabla":
            headers = sec["headers"]
            if headers:
                for j, h in enumerate(headers, 1):
                    c = ws.cell(r, j, h)
                    c.font = white_b
                    c.fill = dark
                r += 1
            for f in sec["filas"]:
                if not isinstance(f, (list, tuple)):
                    continue
                for j, val in enumerate(f, 1):
                    ws.cell(r, j, _xcell(val))
                r += 1
        else:  # kv
            for f in sec["filas"]:
                if not isinstance(f, (list, tuple)) or len(f) < 1:
                    continue
                kc = ws.cell(r, 1, str(f[0]))
                kc.font = label_f
                kc.fill = soft
                if len(f) > 1:
                    ws.cell(r, 2, _xcell(f[1]))
                r += 1
        r += 1

    ws.column_dimensions["A"].width = 34
    for col in ("B", "C", "D", "E", "F", "G"):
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── orquestación: render + almacenamiento + mensaje ──
# Vencimiento del link de descarga/compartir para los entregables financieros.
_SIGNED_URL_TTL = 48 * 3600  # 48 horas


async def _store(filename: str, blob: bytes, content_type: str) -> str | None:
    """Sube a Supabase (signed URL 48h) con fallback a disco. Reusa document_service."""
    try:
        from services.document_service import _save_local, _upload_to_supabase
    except Exception as e:  # noqa: BLE001
        logger.exception("storage helpers no disponibles: %s", e)
        return None
    url = await _upload_to_supabase(filename, blob, content_type, expires_in=_SIGNED_URL_TTL)
    if url:
        return url
    logger.warning(
        "Supabase no disponible: uso fallback de disco (efímero en Railway, no "
        "garantiza 24-48h). Configurá SUPABASE_URL/SERVICE_ROLE_KEY + bucket 'reports'."
    )
    try:
        return _save_local(filename, blob)
    except Exception:  # noqa: BLE001
        logger.exception("fallback local falló")
        return None


async def generar_documento(
    titulo: str | None = None,
    formato: str = "ambos",
    secciones: list | None = None,
    resumen: str | None = None,
    proyecto: str | None = None,
    **_ignore: Any,
) -> dict:
    """Handler del tool `generar_documento_analisis`."""
    if not titulo or not isinstance(secciones, list) or not secciones:
        return {"error": "Necesito 'titulo' y 'secciones' con el contenido del análisis.", "ok": False}

    formato = (formato or "ambos").strip().lower()
    if formato not in ("pdf", "excel", "ambos"):
        formato = "ambos"
    want_pdf = formato in ("pdf", "ambos")
    want_xls = formato in ("excel", "ambos")

    slug = _slug(titulo)
    stamp = datetime.now().strftime("%Y%m%d")
    uid = uuid.uuid4().hex[:8]
    archivos: list[dict] = []
    errores: list[str] = []

    if want_pdf:
        try:
            # to_thread: reportlab es CPU-bound sync; correrlo en el event loop
            # congela TODOS los streams SSE concurrentes mientras genera el PDF.
            blob = await asyncio.to_thread(
                render_analisis_pdf, titulo, secciones, resumen, proyecto
            )
            fn = f"analisis-{slug}-{stamp}-{uid}.pdf"
            url = await _store(fn, blob, "application/pdf")
            if url:
                archivos.append({"formato": "pdf", "filename": fn, "url": url})
            else:
                errores.append("PDF: no se pudo almacenar")
        except Exception as e:  # noqa: BLE001
            logger.exception("render PDF falló")
            errores.append(f"PDF: {e}")

    if want_xls:
        try:
            # to_thread: openpyxl es CPU-bound sync (ver nota del PDF de arriba).
            blob = await asyncio.to_thread(
                render_analisis_xlsx, titulo, secciones, resumen, proyecto
            )
            fn = f"analisis-{slug}-{stamp}-{uid}.xlsx"
            url = await _store(
                fn, blob,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            if url:
                archivos.append({"formato": "excel", "filename": fn, "url": url})
            else:
                errores.append("Excel: no se pudo almacenar")
        except Exception as e:  # noqa: BLE001
            logger.exception("render XLSX falló")
            errores.append(f"Excel: {e}")

    if not archivos:
        return {
            "error": "No pude generar el documento. " + ("; ".join(errores) if errores else ""),
            "ok": False,
        }

    # Mensaje listo para compartir por WhatsApp.
    lineas = [f"*{titulo}*"]
    if resumen:
        lineas.append(resumen)
    for a in archivos:
        etiqueta = "📄 PDF" if a["formato"] == "pdf" else "📊 Excel"
        lineas.append(f"{etiqueta}: {a['url']}")
    lineas.append("— vía RE Expert")
    mensaje = "\n".join(lineas)

    return {
        "ok": True,
        "archivos": archivos,
        "mensaje_whatsapp": mensaje,
        "source": archivos[0]["url"],
        "instruccion": (
            "Mostrá al usuario CADA archivo como un enlace markdown clicleable "
            "([Descargar PDF](url) / [Descargar Excel](url)) y ofrecé el "
            "mensaje_whatsapp para que lo comparta. Los links vencen en 48h."
        ),
        "notas": "; ".join(errores) if errores else None,
    }


# ── schema (formato Anthropic tool_use) ──
DOCUMENT_TOOL_SCHEMA: dict = {
    "name": "generar_documento_analisis",
    "description": (
        "Genera un documento DESCARGABLE (PDF y/o Excel) de un análisis financiero "
        "y devuelve los links + un mensaje listo para WhatsApp. Usalo cuando el "
        "usuario pida bajar/exportar/guardar/mandar el análisis (ej: 'pasámelo en "
        "excel', 'bajámelo en pdf', 'mandámelo por whatsapp', 'quiero el archivo'). "
        "Después de un análisis financiero importante, OFRECELO ('¿te lo bajo en "
        "Excel o PDF?'). Construí `secciones` con el contenido del análisis que ya "
        "hiciste (los mismos números de las tools). IMPORTANTE: pasá los MONTOS como "
        "números CRUDOS (323936, no 'USD 323.936') para que el Excel quede editable; "
        "etiquetas, porcentajes con símbolo y texto van como string. Tras llamarla, "
        "mostrá los links como enlaces markdown clicleables."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "titulo": {"type": "string", "description": "Título del documento, ej 'Factibilidad — Edificio Soler 4500'."},
            "formato": {"type": "string", "enum": ["pdf", "excel", "ambos"], "default": "ambos", "description": "Qué generar. Si el usuario pidió uno puntual, usalo; si no, 'ambos'."},
            "resumen": {"type": "string", "description": "1-2 líneas con el veredicto/headline (va arriba del doc y en el mensaje de WhatsApp)."},
            "proyecto": {"type": "string", "description": "Nombre del proyecto activo, si hay."},
            "secciones": {
                "type": "array",
                "description": "Secciones del documento, en orden.",
                "items": {
                    "type": "object",
                    "properties": {
                        "titulo": {"type": "string", "description": "Título de la sección (ej 'Resultados', 'Flujo por período')."},
                        "tipo": {"type": "string", "enum": ["kv", "tabla"], "description": "'kv' = pares dato/valor (2 columnas); 'tabla' = headers + filas."},
                        "headers": {"type": "array", "items": {"type": "string"}, "description": "Solo para tipo 'tabla': nombres de columna."},
                        "filas": {
                            "type": "array",
                            "description": "Para 'kv': [[etiqueta, valor], ...]. Para 'tabla': [[c1, c2, ...], ...]. Montos como números crudos.",
                            "items": {"type": "array"},
                        },
                    },
                    "required": ["filas"],
                },
            },
        },
        "required": ["titulo", "secciones"],
    },
}
